from openpyxl import load_workbook


def load_excel_kv(path: str, sheet_name: str) -> dict:
    """Author: taobo.zhou
    从 Excel 指定 sheet 读取键值对数据。
    
        path: Excel 文件路径。
        sheet_name: 需要读取的 sheet 名称。
    """

    wb = load_workbook(path, data_only=True)

    if sheet_name not in wb.sheetnames:
        raise RuntimeError(
            f"Excel 中不存在名为 [{sheet_name}] 的 sheet"
        )

    ws = wb[sheet_name]
    data = {}

    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        key, value = row
        if key is None:
            continue
        data[str(key).strip()] = value

    return data


def load_excel_sheets_kv(path: str) -> dict:
    """Author: taobo.zhou
    读取 Excel 中的所有 sheet 并返回键值对集合。
    
        path: Excel 文件路径。
    """

    wb = load_workbook(path, data_only=True)
    if not wb.sheetnames:
        raise RuntimeError("Excel 中至少必须存在一个 sheet")

    result = {}
    for name in wb.sheetnames:
        result[name] = load_excel_kv(path, name)

    return result
