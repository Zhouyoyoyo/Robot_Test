from __future__ import annotations

import json
import os
import sys
import time
import zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional

import subprocess
from openpyxl import load_workbook

from framework.utils.config_loader import load_config
from framework.utils.html_report import build_html_report
from framework.utils.logger import get_logger
from framework.utils.mailer import send_report

log = get_logger()


@dataclass
class CaseResult:
    """Author: taobo.zhou
    中文：用于汇总报告的用例执行结果。
    English: Case result data for report aggregation.
    """

    case_id: str
    sheet: str
    status: str
    retried: bool
    attempt: int
    error: Optional[str]
    screenshot: Optional[str]
    nodeid: str
    start_time: str
    end_time: str


def _now_ts() -> str:
    """Author: taobo.zhou
    中文：获取当前时间戳字符串。
    参数: 无。
    """

    return time.strftime("%Y%m%d_%H%M%S")


def _ensure_dir(path: Path) -> None:
    """Author: taobo.zhou
    中文：确保目录存在。
    参数:
        path: 目标目录路径。
    """

    path.mkdir(parents=True, exist_ok=True)


def _resolve_data_path(cfg: dict) -> Path:
    """Author: taobo.zhou
    中文：解析测试数据文件路径。
    参数:
        cfg: 配置字典。
    """

    project_root = Path(cfg.get("_project_root", "."))
    data_path = Path(cfg["paths"]["data"])
    if not data_path.is_absolute():
        data_path = (project_root / data_path).resolve()
    return data_path


def _load_sheet_names(cfg: dict) -> List[str]:
    """Author: taobo.zhou
    中文：读取 Excel 中所有 sheet 名称。
    参数:
        cfg: 配置字典。
    """

    data_path = _resolve_data_path(cfg)
    if not data_path.exists():
        raise RuntimeError(f"Excel 数据文件不存在: {data_path}")
    wb = load_workbook(str(data_path), read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    if not sheet_names:
        raise RuntimeError("Excel 中至少必须存在一个 sheet")
    return sheet_names


def _run_sheet(sheet: str, run_dir: Path, run_root: Path) -> int:
    """Author: taobo.zhou
    中文：运行指定 sheet 的 pytest 用例。
    参数:
        sheet: sheet 名称。
        run_dir: 当前 sheet 的运行目录。
        run_root: 运行根目录。
    """

    env = os.environ.copy()
    env["PW_WORKER"] = "1"
    env["PW_RUN_DIR"] = str(run_dir)
    env["PW_PINGID_LOCKFILE"] = str(run_root / "pingid.lock")

    cmd = [
        sys.executable,
        "-m",
        "pytest",
        "-q",
        "--pw-worker",
        "--pw-sheet",
        sheet,
        "--pw-run-dir",
        str(run_dir),
    ]
    log.info("[PW][RUN] %s", " ".join(cmd))
    completed = subprocess.run(cmd, env=env)
    return completed.returncode


def _normalize_status(status: str) -> str:
    """Author: taobo.zhou
    中文：规范化状态字符串用于报告输出。
    参数:
        status: 原始状态字符串。
    """

    mapping = {
        "PASSED": "PASS",
        "FAILED": "FAIL",
        "SKIPPED": "SKIP",
        "PASS": "PASS",
        "FAIL": "FAIL",
        "SKIP": "SKIP",
        "ERROR": "ERROR",
    }
    return mapping.get(status, status)


def _collect_results(run_root: Path, sheet_names: List[str]) -> tuple[List[CaseResult], Dict[str, dict], Dict[str, int]]:
    """Author: taobo.zhou
    中文：汇总所有 sheet 的运行结果。
    参数:
        run_root: 运行根目录。
        sheet_names: sheet 名称列表。
    """

    results: List[CaseResult] = []
    case_params: Dict[str, dict] = {}
    counts = {"total": 0, "passed": 0, "failed": 0, "error": 0, "skipped": 0}

    for sheet in sheet_names:
        result_path = run_root / sheet / "reports" / "results.json"
        if not result_path.exists():
            log.error("[PW][SUMMARY] missing results.json for sheet=%s", sheet)
            counts["error"] += 1
            counts["total"] += 1
            continue

        with result_path.open("r", encoding="utf-8") as f:
            payload = json.load(f)

        case_params.update(payload.get("case_params", {}) or {})
        for item in payload.get("results", []):
            status = _normalize_status(str(item.get("status", "")))
            results.append(CaseResult(
                case_id=str(item.get("case_id", "")),
                sheet=str(item.get("sheet", "")),
                status=status,
                retried=bool(item.get("retried")),
                attempt=int(item.get("attempt") or 1),
                error=item.get("error"),
                screenshot=item.get("screenshot"),
                nodeid=str(item.get("nodeid", "")),
                start_time=str(item.get("start_time", "-")),
                end_time=str(item.get("end_time", "-")),
            ))

            counts["total"] += 1
            if status == "PASS":
                counts["passed"] += 1
            elif status == "FAIL":
                counts["failed"] += 1
            elif status == "ERROR":
                counts["error"] += 1
            elif status == "SKIP":
                counts["skipped"] += 1

    return results, case_params, counts


def _zip_screenshots(run_root: Path, ts: str) -> Optional[str]:
    """Author: taobo.zhou
    中文：打包所有截图并返回 zip 路径。
    参数:
        run_root: 运行根目录。
        ts: 时间戳字符串。
    """

    reports_dir = run_root / "reports"
    _ensure_dir(reports_dir)
    zip_path = reports_dir / f"screenshots_{ts}.zip"
    with zipfile.ZipFile(str(zip_path), "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for sheet_dir in run_root.iterdir():
            if not sheet_dir.is_dir():
                continue
            ss_dir = sheet_dir / "screenshots"
            if not ss_dir.exists():
                continue
            for root, _, files in os.walk(str(ss_dir)):
                for fname in files:
                    file_path = Path(root) / fname
                    rel_path = file_path.relative_to(run_root)
                    zf.write(str(file_path), str(rel_path))
    return str(zip_path)


def main() -> int:
    """Author: taobo.zhou
    中文：主入口，执行并汇总自动化测试。
    参数: 无。
    """

    cfg = load_config()
    max_workers = int(cfg.get("runner", {}).get("max_workers", 1))
    sheet_names = _load_sheet_names(cfg)

    ts = _now_ts()
    project_root = Path(cfg.get("_project_root", "."))
    run_root = project_root / "output" / "runs" / ts
    _ensure_dir(run_root)

    for sheet in sheet_names:
        sheet_dir = run_root / sheet
        _ensure_dir(sheet_dir / "screenshots")
        _ensure_dir(sheet_dir / "reports")

    returncodes = {}
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {
            executor.submit(_run_sheet, sheet, run_root / sheet, run_root): sheet
            for sheet in sheet_names
        }
        for future in as_completed(futures):
            sheet = futures[future]
            try:
                returncodes[sheet] = future.result()
            except Exception as exc:
                log.error("[PW][RUN] sheet=%s failed: %s", sheet, exc)
                returncodes[sheet] = 1

    results, case_params, counts = _collect_results(run_root, sheet_names)

    reports_dir = run_root / "reports"
    _ensure_dir(reports_dir)
    report_path = reports_dir / f"report_{ts}.html"
    report_info = build_html_report(results, case_params)
    html = report_info.get("html") or report_info.get("html_content") or report_info.get("content")
    if not html:
        html = (
            "<html><body><h2>Report</h2>"
            f"<p>Total={counts['total']} Pass={counts['passed']} "
            f"Fail={counts['failed']} Error={counts['error']} Skip={counts['skipped']}</p>"
            "</body></html>"
        )
    report_path.write_text(html, encoding="utf-8")

    screenshot_zip = _zip_screenshots(run_root, ts)
    subject = (
        f"Robot 自动化测试报告 | Total={counts['total']} "
        f"Pass={counts['passed']} Fail={counts['failed']} "
        f"Error={counts['error']} Skip={counts['skipped']}"
    )
    details = [
        {
            "sheet": r.sheet,
            "status": r.status,
            "attempt": r.attempt,
            "error": r.error,
            "screenshot": r.screenshot,
        }
        for r in results
    ]
    send_report(
        pytest_results={
            "total": counts["total"],
            "passed": counts["passed"],
            "failed": counts["failed"],
            "error": counts["error"],
            "skipped": counts["skipped"],
            "details": details,
        },
        html_report=str(report_path),
        screenshot_zip=screenshot_zip,
        subject=subject,
        extra_attachments=None,
    )

    if counts["failed"] > 0 or counts["error"] > 0:
        return 1
    if any(code != 0 for code in returncodes.values()):
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
