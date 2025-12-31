from pathlib import Path

import pytest
from openpyxl import load_workbook

from framework.core.driver_manager import DriverManager
from framework.utils.config_loader import load_config
from framework.utils.excel_loader import load_excel_kv
from framework.utils.locator_loader import LocatorLoader


@pytest.fixture(scope="session")
def config():
    """Author: taobo.zhou
    加载并补全全局配置。
     无。
    """

    cfg = load_config()

    project_root = Path(cfg.get("_project_root", "."))
    if "paths" in cfg and isinstance(cfg["paths"], dict):
        for k, v in list(cfg["paths"].items()):
            if isinstance(v, str) and v and not Path(v).is_absolute():
                cfg["paths"][k] = str((project_root / v).resolve())

    locator_path = cfg["paths"]["locator"]
    loader = LocatorLoader(locator_path)
    loader.validate_all()
    cfg["locator_loader"] = loader

    return cfg


@pytest.fixture(scope="session")
def driver(config, request):
    """Author: taobo.zhou
    初始化浏览器驱动并在会话结束时关闭。
    
        config: 全局配置字典。
        request: pytest 请求对象，用于挂载 driver。
    """

    driver = DriverManager.get_driver()
    request.session.driver = driver

    selenium_cfg = config.get("selenium", {}) or {}
    implicit_wait = selenium_cfg.get("implicit_wait")
    if implicit_wait is not None:
        try:
            driver.implicitly_wait(float(implicit_wait))
        except Exception:
            pass

    page_load_timeout = selenium_cfg.get("page_load_timeout")
    if page_load_timeout is not None:
        try:
            driver.set_page_load_timeout(float(page_load_timeout))
        except Exception:
            pass

    yield driver
    DriverManager.quit()


def pytest_generate_tests(metafunc):
    """Author: taobo.zhou
    将 Excel 中的每个 sheet 转换为 pytest 用例。
    
        metafunc: pytest 的参数化元对象。
    """

    if "sheet_name" not in metafunc.fixturenames:
        return

    cfg = load_config()

    project_root = Path(cfg.get("_project_root", "."))
    data_path = Path(cfg["paths"]["data"])
    if not data_path.is_absolute():
        data_path = (project_root / data_path).resolve()

    if not data_path.exists():
        raise RuntimeError(f"Excel 数据文件不存在: {data_path}")

    wb = load_workbook(str(data_path), read_only=True, data_only=True)
    sheet_names = wb.sheetnames

    if not sheet_names:
        raise RuntimeError("Excel 中至少必须存在一个 sheet")

    selected_sheet = metafunc.config.getoption("--pw-sheet")
    if selected_sheet:
        if selected_sheet not in sheet_names:
            raise RuntimeError(f"指定的 sheet 不存在: {selected_sheet}")
        sheet_names = [selected_sheet]

    metafunc.parametrize(
        "sheet_name",
        sheet_names,
        ids=[f"sheet={name}" for name in sheet_names],
    )


@pytest.fixture
def case_data(config, sheet_name):
    """Author: taobo.zhou
    读取当前 sheet_name 对应的测试数据。
    
        config: 全局配置字典。
        sheet_name: Excel sheet 名称。
    """

    return load_excel_kv(
        config["paths"]["data"],
        sheet_name,
    )


@pytest.fixture
def base_url(config, sheet_name):
    """Author: taobo.zhou
    获取当前 sheet_name 对应的基础 URL。
    
        config: 全局配置字典。
        sheet_name: Excel sheet 名称。
    """

    urls = config.get("project", {}).get("urls")
    if not isinstance(urls, dict):
        raise RuntimeError("config.yaml 中缺少 project.urls")

    if sheet_name not in urls:
        raise RuntimeError(
            f"config.yaml 的 project.urls 中未配置 sheet [{sheet_name}] 的 URL"
        )

    return urls[sheet_name]


pytest_plugins = ["tests.pytest_hooks"]
